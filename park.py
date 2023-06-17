import cv2
import imutils
import pytesseract
import openpyxl
import os
import glob
import sys
import datetime
from openpyxl.utils import coordinate_to_tuple
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers
import getpass
import re
import time
import numpy as np
import pygame
from PIL import Image

parking_space=[]
for i in range (0,10):
    temp=[]
    for j in range(0,10):
        temp.append(0);
    parking_space.append(temp)


def aadhar():
    pytesseract.pytesseract.tesseract_cmd = 'C:\python\Lib\site-packages\Tesseract-OCR\\tesseract'

    dir_path = r"C:\python project\boombabys\aadhar"

    # Get a list of all files in the directory with the .jpg extension
    file_list = glob.glob(os.path.join(dir_path, "*.jpg"))

    # Sort the list of files by date modified (most recent first)
    file_list.sort(key=lambda x: os.path.getmtime(x), reverse=True)

    # Open the most recent file using PIL
    most_recent_file = file_list[0]
    image = Image.open(most_recent_file)

    # Display the image
    image.show()
    #img3 = cv2.imread('test_aadhar2.jpg')  # any picture
    img3 = cv2.imread(image.filename)
    print(img3.shape)

    # removing shadow/noise from image which can be taken from phone camera

    rgb_planes = cv2.split(img3)
    result_planes = []
    result_norm_planes = []
    for plane in rgb_planes:
        dilated_img = cv2.dilate(plane,
                                 np.ones((10, 10), np.uint8))  # change the value of (10,10) to see different results
        bg_img = cv2.medianBlur(dilated_img, 21)
        diff_img = 255 - cv2.absdiff(plane, bg_img)
        norm_img = cv2.normalize(diff_img, None, alpha=0, beta=250, norm_type=cv2.NORM_MINMAX,
                                 dtype=cv2.CV_8UC1)
        result_planes.append(diff_img)
        result_norm_planes.append(norm_img)

    result = cv2.merge(result_planes)
    result_norm = cv2.merge(result_norm_planes)
    dst = cv2.fastNlMeansDenoisingColored(result_norm, None, 10, 10, 7, 11)  # removing noise from image

    text = pytesseract.image_to_string(dst).upper().replace(" ", "")

    date = str(re.findall(r"[\d]{1,4}[/-][\d]{1,4}[/-][\d]{1,4}", text)).replace("]", "").replace("[", "").replace("'",
                                                                                                                   "")
    print(date)
    number = str(re.findall(r"[0-9]{11,12}", text)).replace("]", "").replace("[", "").replace("'", "")
    print(number)
    sex = str(re.findall(r"MALE|FEMALE", text)).replace("[", "").replace("'", "").replace("]", "")
    print(sex)

    return [number,date]

pygame.mixer.init()
pygame.mixer.music.load('kola.mp3')
pygame.mixer.music.play(-1)
while(1):
    print("Welcome to X parking system")
    print("please choose the suitable category:")

    print("1. User/Guest")
    print("2.Guard")
    print("3.Admin")
    option = int(input("please enter the desired option from above categories(1/2/3)"))

    if (option == 1):

        pytesseract.pytesseract.tesseract_cmd = 'C:\python\Lib\site-packages\Tesseract-OCR\\tesseract'

        dir_path = r"C:\python project\boombabys\car"

        # Get a list of all files in the directory with the .jpg extension
        file_list = glob.glob(os.path.join(dir_path, "*.jpg"))

        # Sort the list of files by date modified (most recent first)
        file_list.sort(key=lambda x: os.path.getmtime(x), reverse=True)

        # Open the most recent file using PIL
        most_recent_file = file_list[0]
        image = Image.open(most_recent_file)

        # Display the image
        image.show()

        print("please enter the suitable option")
        print("1.Entry")
        print("2.Exit")
        option_user = int(input("(1/2)?"))

        image = cv2.imread(image.filename)
        image = imutils.resize(image, width=300)
        # cv2.imshow("original image", image)
        # cv2.waitKey(0)
        gray_image = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        # cv2.imshow("greyed image", gray_image)
        # cv2.waitKey(0)
        gray_image = cv2.bilateralFilter(gray_image, 11, 17, 17)
        # cv2.imshow("smoothened image", gray_image)
        # cv2.waitKey(0)
        edged = cv2.Canny(gray_image, 30, 200)
        # cv2.imshow("edged image", edged)
        # cv2.waitKey(0)
        cnts, new = cv2.findContours(edged.copy(), cv2.RETR_LIST, cv2.CHAIN_APPROX_SIMPLE)
        image1 = image.copy()
        cv2.drawContours(image1, cnts, -1, (0, 255, 0), 3)
        # cv2.imshow("contours",image1)
        # cv2.waitKey(0)
        cnts = sorted(cnts, key=cv2.contourArea, reverse=True)[:30]
        screenCnt = None
        image2 = image.copy()
        cv2.drawContours(image2, cnts, -1, (0, 255, 0), 3)
        # cv2.imshow("Top 30 contours",image2)
        # cv2.waitKey(0)
        i = 7
        for c in cnts:
            perimeter = cv2.arcLength(c, True)
            approx = cv2.approxPolyDP(c, 0.018 * perimeter, True)
            if len(approx) == 4:
                screenCnt = approx
                x, y, w, h = cv2.boundingRect(c)
                new_img = image[y:y + h, x:x + w]
                cv2.imwrite('./' + str(i) + '.jpeg', new_img)
                i += 1
                break
        cv2.drawContours(image, [screenCnt], -1, (0, 255, 0), 3)
        # cv2.imshow("image with detected license plate", image)
        # cv2.waitKey(0)
        Cropped_loc = './7.jpeg'
        # cv2.imshow("cropped", cv2.imread(Cropped_loc))
        plate = pytesseract.image_to_string(Cropped_loc, lang='eng')
        print("Number plate is:", plate)
        # cv2.waitKey(0)
        # cv2.destroyAllWindows()

        if (option_user == 1):
            print("Please provide aadhar card");
            temp=aadhar()
            number=temp[0];
            date=temp[1];
            print(number);
            space = ""
            found = False
            for i in range(0, 10):
                for j in range(0, 10):
                    if (parking_space[i][j] == 0):
                        parking_space[i][j] = 1
                        space = str(i);
                        space += str(j);
                        found = True;
                        break;
                if (found):
                    break;

            if os.path.exists('my_database.xlsx'):
                # Load the existing workbook
                workbook = openpyxl.load_workbook('my_database.xlsx')
                worksheet = workbook.active
                column = worksheet['C']
                search_value = plate

                for cell in column[0:]:
                    if cell.value == search_value:
                        # The value was found
                        row = coordinate_to_tuple(cell.coordinate)[1]
                        now = datetime.datetime.now()
                        new_values = [number,date, plate, now, -1, 0,space]

                        for i, value in enumerate(new_values):
                            worksheet.cell(row=row, column=i + 1, value=value)

                        break
                else:
                    # The value was not found
                    now = datetime.datetime.now()
                    worksheet.append([number,date, plate, now, -1, 0,space])

                # Save the changes to the workbook
                workbook.save('my_database.xlsx')
            else:
                # Create a new workbook
                workbook = openpyxl.Workbook()
                # Select the first worksheet
                worksheet = workbook.active
                # Add some data to the worksheet
                worksheet['A1'] = 'Aadhar'
                worksheet['B1'] = 'D.O.B'
                worksheet['C1'] = 'License Plate'
                worksheet['D1'] = 'In_time'
                worksheet['E1'] = 'Out_time'
                worksheet['F1'] = 'Amount'
                worksheet['G1'] = 'Position'
                # Set the width of all cells in column A to fit the date and time format
                column_letter = openpyxl.utils.get_column_letter(4)
                column_dimensions = worksheet.column_dimensions[column_letter]
                column_dimensions.width = len('MM/DD/YYYY HH:MM:SS') + 2  # Add 2 for padding

                # Set the number format for each cell in the column

                column_letter = get_column_letter(4)
                for row in worksheet.iter_rows(min_row=1, max_col=1, max_row=worksheet.max_row):
                    cell = row[0]
                    if cell.column_letter == column_letter:
                        cell.number_format = numbers.FORMAT_DATE_DATETIME

                # Set the width of all cells in column A to fit the date and time format
                column_letter = openpyxl.utils.get_column_letter(5)
                column_dimensions = worksheet.column_dimensions[column_letter]
                column_dimensions.width = len('MM/DD/YYYY HH:MM:SS') + 2  # Add 2 for padding
                # Set the number format for each cell in the column
                column_letter = get_column_letter(5)
                for row in worksheet.iter_rows(min_row=1, max_col=1, max_row=worksheet.max_row):
                    cell = row[0]
                    if cell.column_letter == column_letter:
                        cell.number_format = numbers.FORMAT_DATE_DATETIME

                now = datetime.datetime.now()
                worksheet.append([number,date, plate, now, now, 0,space])
                # Save the workbook to a new file
                workbook.save('my_database.xlsx')

        else:
            if os.path.exists('my_database.xlsx'):
                # Load the existing workbook
                workbook = openpyxl.load_workbook('my_database.xlsx')
                worksheet = workbook.active
                column = worksheet['C']
                search_value = plate

                for cell in column[0:]:
                    if cell.value == search_value:
                        # The value was found
                        row = coordinate_to_tuple(cell.coordinate)[1]
                        number = worksheet['A' + str(row)].value
                        date = worksheet['B' + str(row)].value
                        space=worksheet['G'+str(row)].value
                        parking_space[int(space[0])][int(space[1])] = 0;
                        in_time = worksheet.cell(row=row, column=4).value

                        if in_time is not None:
                            datetime_value = in_time.date()

                            # Print the datetime value
                        else:
                            print("please call support")

                        out_time = datetime.datetime.now()
                        amount = (((out_time - in_time).total_seconds()) * (0.0005))
                        new_values = [number,date, plate, in_time, out_time, amount]
                        print("Amount to be paid is:" + str(amount))
                        for i, value in enumerate(new_values):
                            worksheet.cell(row=row, column=i + 1, value=value)

                        break
                else:
                    # The value was not found
                    print("Details not found..Please call the guard")

                # Save the changes to the workbook
                workbook.save('my_database.xlsx')
            else:
                # no data
                print("Details not found..Please call the guard")

    elif (option == 2):
        plate = input("Please enter the number plate")
        if os.path.exists('my_database.xlsx'):
            # Load the existing workbook
            workbook = openpyxl.load_workbook('my_database.xlsx')
            worksheet = workbook.active
            column = worksheet['C']
            search_value = plate
            for cell in column[1:]:
                cell.value = cell.value.strip()

                if cell.value == search_value:
                    # The value was found
                    row = coordinate_to_tuple(cell.coordinate)[1]
                    name = worksheet['A' + str(row)].value
                    phone = worksheet['B' + str(row)].value
                    space = worksheet['G' + str(row)].value
                    in_time = worksheet.cell(row=row, column=4).value

                    if in_time is not None:
                        datetime_value = in_time.date()

                        # Print the datetime value
                    else:
                        print("please call support")

                    out_time = datetime.datetime.now()
                    amount = (((out_time - in_time).total_seconds()) * (0.0005))
                    new_values = [name, phone, plate, in_time, out_time, amount,space]
                    print("Details are as follows:")
                    print("Aadhar:" + str(new_values[0]))
                    print("D.O.B:" + str(new_values[1]))
                    print("Plate:" + str(new_values[2]))
                    print("Position:" +str(chr(ord('A')+int(new_values[6][0]))+str(new_values[6][1])))
                    break
            else:
                # The value was not found
                print("Details not found..Please call the support")

            # Save the changes to the workbook

        else:
            # no data
            print("Details not found..Please call the guard")
    elif (option == 3):
        password = getpass.getpass('Enter the password:')
        if password == 'admin123':
            print("Please select the option you want:")
            print("1.Show database")
            print("2.Terminate application")
            n = int(input())
            if (n == 1):
                excel_file_path = 'my_database.xlsx'

                # Define the command to open the Excel file in Microsoft Excel
                excel_command = 'start excel.exe "{}"'.format(excel_file_path)

                # Execute the command to open the Excel file
                os.system(excel_command)
            elif (n == 2):
                sys.exit(0)

        else:
            print("Wrong password")
    time.sleep(5)
    if os.name == 'nt':
        _ = os.system('cls')

    # For Mac and Linux
    else:
        _ = os.system('clear')














